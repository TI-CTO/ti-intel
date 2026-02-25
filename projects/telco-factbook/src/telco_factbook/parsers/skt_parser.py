"""SKT Excel financial statement parser."""

from __future__ import annotations

import logging
import re
from pathlib import Path

import openpyxl

from telco_factbook.models import (
    Carrier,
    FinancialMetrics,
    ParseIssue,
    ParseResult,
    PeriodType,
)
from telco_factbook.parsers.base import BaseExcelParser

logger = logging.getLogger(__name__)

# Known field name patterns in SKT English Excel files
_FIELD_PATTERNS: dict[str, list[str]] = {
    "revenue": ["Operating revenue", "Total Revenue", "Operating Revenue", "매출액"],
    "operating_income": ["Operating income", "Operating Profit", "영업이익"],
    "net_income": ["Net income", "Net Profit", "당기순이익"],
    "ebitda": ["EBITDA"],
    "revenue_mobile": ["Mobile service revenue", "MNO"],
    "revenue_fixed": ["Fixed-line", "Wireline", "유선"],
    "revenue_media": ["Media", "Content", "미디어"],
    "revenue_enterprise": ["Enterprise", "B2B", "기업"],
    "total_assets": ["Total assets", "자산총계"],
    "total_debt": ["Total liabilities", "Total Debt", "부채총계"],
    "capex": ["CAPEX", "Capital Expenditure", "설비투자"],
    "mobile_subscribers": ["Mobile Subscribers", "Wireless Subscribers", "가입자"],
    "mobile_5g_subscribers": ["5G Subscribers", "5G 가입자"],
}

# Quarter to BS date-header month mapping
_QUARTER_MONTH: dict[int, str] = {1: "3-31", 2: "6-30", 3: "9-30", 4: "12-31"}


class SKTExcelParser(BaseExcelParser):
    """Parser for SKT quarterly financial statement Excel files."""

    carrier = Carrier.SKT

    def parse(self, file_path: Path, year: int, quarter: int) -> ParseResult:
        """Parse SKT Excel financial statement."""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        issues: list[ParseIssue] = []

        metrics_dict: dict[str, int | float | None] = {}
        sheets_checked = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            found = self._scan_sheet(ws, year, quarter, metrics_dict, issues)
            sheets_checked += 1
            if found:
                logger.info("Found data in sheet: %s", sheet_name)

        wb.close()

        key_fields = ["revenue", "operating_income", "net_income"]
        found_keys = sum(1 for f in key_fields if metrics_dict.get(f) is not None)
        confidence = found_keys / len(key_fields)

        if confidence == 0:
            issues.append(
                ParseIssue(
                    field_name="all",
                    issue_type="missing",
                    description=f"No financial data found in {sheets_checked} sheets for {year}Q{quarter}",
                )
            )

        metrics = FinancialMetrics(
            carrier=Carrier.SKT,
            year=year,
            quarter=quarter,
            period_type=PeriodType.QUARTERLY,
            **metrics_dict,
        )

        return ParseResult(metrics=metrics, confidence=confidence, issues=issues)

    def _scan_sheet(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        year: int,
        quarter: int,
        metrics_dict: dict[str, int | float | None],
        issues: list[ParseIssue],
    ) -> bool:
        """Scan a worksheet for financial data rows."""
        found_any = False
        header_row = self._detect_header_row(ws)
        quarter_col = self._find_quarter_column(ws, header_row, year, quarter) if header_row else None

        if quarter_col is None:
            return False

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            label_cell = row[0] if row else None
            if label_cell is None or label_cell.value is None:
                continue

            label = str(label_cell.value).strip()

            for field_name, patterns in _FIELD_PATTERNS.items():
                if field_name in metrics_dict:
                    continue
                if any(p.lower() == label.lower() or label.lower().startswith(p.lower()) for p in patterns):
                    value = None
                    if quarter_col < len(row):
                        value = row[quarter_col].value

                    if value is not None:
                        converted = self._safe_int(value)
                        if converted is not None:
                            metrics_dict[field_name] = converted
                            found_any = True
                        else:
                            issues.append(
                                ParseIssue(
                                    field_name=field_name,
                                    raw_value=str(value),
                                    issue_type="format_mismatch",
                                    description=f"Could not convert '{value}' to int",
                                )
                            )
                    break

        return found_any

    def _detect_header_row(
        self, ws: openpyxl.worksheet.worksheet.Worksheet
    ) -> int | None:
        """Find the row containing quarter headers (1Q24, 3-31-24, etc.)."""
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
            for cell_val in row:
                if cell_val is None:
                    continue
                cell_str = str(cell_val).upper()
                # IS sheets: "1Q24", "2Q25" etc.
                if re.search(r"\dQ\d{2}", cell_str):
                    return row_idx
                # BS sheets: "3-31-24", "12-31-25" etc.
                if re.search(r"\d{1,2}-\d{1,2}-\d{2}", cell_str):
                    return row_idx
        return None

    def _find_quarter_column(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        header_row: int,
        year: int,
        quarter: int,
    ) -> int | None:
        """Find column index for a specific year+quarter in the header row.

        Handles two header formats:
        - IS sheets: "1Q24", "4Q25" (quarter + 2-digit year)
        - BS sheets: "3-31-24", "12-31-25" (month-day-2digit year)
        """
        year_suffix = str(year)[-2:]  # "2025" → "25"

        # IS format: "{quarter}Q{year_suffix}"
        is_target = f"{quarter}Q{year_suffix}"

        # BS format: "{month}-{day}-{year_suffix}"
        bs_month = _QUARTER_MONTH.get(quarter, "")
        bs_target = f"{bs_month}-{year_suffix}" if bs_month else ""

        for col_idx, cell in enumerate(ws[header_row]):
            if cell.value is None:
                continue
            cell_str = str(cell.value).strip().upper()

            if is_target.upper() in cell_str:
                return col_idx
            if bs_target and bs_target.upper() in cell_str:
                return col_idx

        return None
