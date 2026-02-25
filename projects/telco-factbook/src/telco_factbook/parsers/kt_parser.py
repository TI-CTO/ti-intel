"""KT Excel financial statement parser."""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl

from telco_factbook.models import (
    Carrier,
    FinancialMetrics,
    ParseIssue,
    ParseResult,
    PeriodType,
)
from telco_factbook.parsers.base import BaseExcelParser

logger = logging.getLogger(__name__)

# Known field name patterns in KT Excel files (Korean)
_FIELD_PATTERNS: dict[str, list[str]] = {
    "revenue": ["매출액", "영업수익", "총매출", "Revenue"],
    "operating_income": ["영업이익", "Operating Income", "Operating Profit"],
    "net_income": ["당기순이익", "순이익", "Net Income", "Net Profit"],
    "ebitda": ["EBITDA"],
    "revenue_mobile": ["무선", "이동전화", "Mobile"],
    "revenue_fixed": ["유선", "초고속인터넷", "Fixed"],
    "revenue_media": ["미디어", "콘텐츠", "Media"],
    "revenue_enterprise": ["기업", "B2B", "Enterprise"],
    "total_assets": ["자산총계", "Total Assets"],
    "total_debt": ["부채총계", "Total Liabilities", "Total Debt"],
    "capex": ["설비투자", "CAPEX", "Capital Expenditure"],
    "mobile_subscribers": ["이동전화가입자", "무선가입자", "Mobile Subscribers"],
}


class KTExcelParser(BaseExcelParser):
    """Parser for KT quarterly financial statement Excel files."""

    carrier = Carrier.KT

    def parse(self, file_path: Path, year: int, quarter: int) -> ParseResult:
        """Parse KT Excel financial statement."""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        issues: list[ParseIssue] = []

        metrics_dict: dict[str, int | float | None] = {}
        sheets_checked = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            found = self._scan_sheet(ws, quarter, metrics_dict, issues)
            sheets_checked += 1
            if found:
                logger.info("Found data in sheet: %s", sheet_name)

        wb.close()

        # Calculate confidence based on how many key fields were found
        key_fields = ["revenue", "operating_income", "net_income"]
        found_keys = sum(1 for f in key_fields if metrics_dict.get(f) is not None)
        confidence = found_keys / len(key_fields)

        if confidence == 0:
            issues.append(
                ParseIssue(
                    field_name="all",
                    issue_type="missing",
                    description=f"No financial data found in {sheets_checked} sheets",
                )
            )

        metrics = FinancialMetrics(
            carrier=Carrier.KT,
            year=year,
            quarter=quarter,
            period_type=PeriodType.QUARTERLY,
            **metrics_dict,
        )

        return ParseResult(metrics=metrics, confidence=confidence, issues=issues)

    def _scan_sheet(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        quarter: int,
        metrics_dict: dict[str, int | float | None],
        issues: list[ParseIssue],
    ) -> bool:
        """Scan a worksheet for financial data rows."""
        found_any = False

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            label_cell = row[0] if row else None
            if label_cell is None or label_cell.value is None:
                continue

            label = str(label_cell.value).strip()

            for field_name, patterns in _FIELD_PATTERNS.items():
                if any(p in label for p in patterns):
                    # Try to find the value in subsequent columns
                    # Quarter columns are typically ordered: 1Q, 2Q, 3Q, 4Q
                    value = self._find_quarter_value(row, quarter)
                    if value is not None:
                        converted = self._safe_int(value)
                        if converted is not None:
                            metrics_dict[field_name] = converted
                            found_any = True
                        else:
                            issues.append(
                                ParseIssue(
                                    field_name=field_name,
                                    raw_value=str(value),
                                    issue_type="format_mismatch",
                                    description=f"Could not convert '{value}' to int",
                                )
                            )
                    break

        return found_any

    def _find_quarter_value(
        self, row: tuple, quarter: int
    ) -> object | None:
        """Find the value for a specific quarter in a row.

        Excel files may have different column layouts. This tries common patterns:
        - Column index offset (label in col 0, Q1 in col 1, Q2 in col 2, etc.)
        - Header-based detection (look for Q1/Q2/Q3/Q4 in header row)
        """
        # Simple approach: assume quarter data starts at column 1
        # and columns are in order Q1, Q2, Q3, Q4
        target_col = quarter  # Q1=col1, Q2=col2, etc.
        if target_col < len(row) and row[target_col].value is not None:
            return row[target_col].value

        # Fallback: try all columns after label
        for cell in row[1:]:
            if cell.value is not None:
                return cell.value

        return None
