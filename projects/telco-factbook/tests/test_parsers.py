"""Tests for KT and SKT Excel parser logic (no real Excel files required)."""

from __future__ import annotations

from unittest.mock import MagicMock, PropertyMock

import openpyxl
import pytest

from telco_factbook.parsers.base import BaseExcelParser
from telco_factbook.parsers.kt_parser import KTExcelParser
from telco_factbook.parsers.skt_parser import SKTExcelParser


# ---------------------------------------------------------------------------
# Concrete subclass for testing BaseExcelParser abstract methods
# ---------------------------------------------------------------------------


class _ConcreteParser(BaseExcelParser):
    """Minimal concrete implementation to exercise BaseExcelParser methods."""

    from telco_factbook.models import Carrier

    carrier = Carrier.KT

    def parse(self, file_path, year, quarter):
        raise NotImplementedError


# ---------------------------------------------------------------------------
# BaseExcelParser._safe_int
# ---------------------------------------------------------------------------


class TestSafeInt:
    def setup_method(self):
        self.parser = _ConcreteParser()

    def test_integer_value(self):
        assert self.parser._safe_int(12345) == 12345

    def test_float_value_truncated(self):
        assert self.parser._safe_int(12345.9) == 12345

    def test_string_integer(self):
        assert self.parser._safe_int("9876") == 9876

    def test_comma_separated_number(self):
        assert self.parser._safe_int("1,234,567") == 1234567

    def test_number_with_spaces(self):
        assert self.parser._safe_int("  42 ") == 42

    def test_none_returns_none(self):
        assert self.parser._safe_int(None) is None

    def test_non_numeric_string_returns_none(self):
        assert self.parser._safe_int("N/A") is None

    def test_empty_string_returns_none(self):
        assert self.parser._safe_int("") is None

    def test_zero(self):
        assert self.parser._safe_int(0) == 0

    def test_negative_number(self):
        assert self.parser._safe_int(-500) == -500


# ---------------------------------------------------------------------------
# BaseExcelParser._safe_float
# ---------------------------------------------------------------------------


class TestSafeFloat:
    def setup_method(self):
        self.parser = _ConcreteParser()

    def test_float_value(self):
        assert self.parser._safe_float(3.14) == pytest.approx(3.14)

    def test_integer_value(self):
        assert self.parser._safe_float(5) == pytest.approx(5.0)

    def test_percentage_string(self):
        assert self.parser._safe_float("12.5%") == pytest.approx(12.5)

    def test_comma_separated(self):
        assert self.parser._safe_float("1,234.5") == pytest.approx(1234.5)

    def test_none_returns_none(self):
        assert self.parser._safe_float(None) is None

    def test_non_numeric_returns_none(self):
        assert self.parser._safe_float("abc") is None


# ---------------------------------------------------------------------------
# KTExcelParser._find_quarter_value
# ---------------------------------------------------------------------------


def _make_row(*values) -> tuple:
    """Build a tuple of mock cells with .value attributes."""
    cells = []
    for v in values:
        cell = MagicMock()
        cell.value = v
        cells.append(cell)
    return tuple(cells)


class TestKTFindQuarterValue:
    def setup_method(self):
        self.parser = KTExcelParser()

    def test_returns_target_column_value(self):
        # Q2 → column index 2 (0-indexed: label, Q1, Q2, Q3, Q4)
        row = _make_row("매출액", 100, 200, 300, 400)
        assert self.parser._find_quarter_value(row, quarter=2) == 200

    def test_q1_returns_first_data_column(self):
        row = _make_row("영업이익", 111, 222, 333, 444)
        assert self.parser._find_quarter_value(row, quarter=1) == 111

    def test_q4_returns_fourth_data_column(self):
        row = _make_row("순이익", 10, 20, 30, 40)
        assert self.parser._find_quarter_value(row, quarter=4) == 40

    def test_target_column_none_falls_back(self):
        """When target column is None, use the first non-None column after label."""
        row = _make_row("EBITDA", None, None, 999)
        # Q2 → index 2 which is None; fallback finds 999 at index 3
        result = self.parser._find_quarter_value(row, quarter=2)
        assert result == 999

    def test_all_none_returns_none(self):
        row = _make_row("자산총계", None, None, None, None)
        assert self.parser._find_quarter_value(row, quarter=1) is None

    def test_row_shorter_than_target_falls_back(self):
        """Row has only label; no data column → fallback also finds nothing."""
        row = _make_row("매출액")
        assert self.parser._find_quarter_value(row, quarter=3) is None


# ---------------------------------------------------------------------------
# SKTExcelParser._detect_header_row (using real openpyxl Workbook)
# ---------------------------------------------------------------------------


def _make_ws_with_header(header_values: list, row: int = 1):
    """Create an openpyxl worksheet with given values in a specific row."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for col, val in enumerate(header_values, start=1):
        ws.cell(row=row, column=col, value=val)
    return ws


class TestSKTDetectHeaderRow:
    def setup_method(self):
        self.parser = SKTExcelParser()

    def test_is_format_detected(self):
        ws = _make_ws_with_header(["Item", "1Q24", "2Q24", "3Q24", "4Q24"], row=2)
        result = self.parser._detect_header_row(ws)
        assert result == 2

    def test_bs_format_detected(self):
        ws = _make_ws_with_header(["Item", "3-31-24", "6-30-24"], row=3)
        result = self.parser._detect_header_row(ws)
        assert result == 3

    def test_no_header_returns_none(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        for r in range(1, 5):
            ws.cell(row=r, column=1, value=f"Row {r} data")
        result = self.parser._detect_header_row(ws)
        assert result is None

    def test_header_in_first_row(self):
        ws = _make_ws_with_header(["Label", "1Q25", "2Q25"], row=1)
        result = self.parser._detect_header_row(ws)
        assert result == 1


# ---------------------------------------------------------------------------
# SKTExcelParser._find_quarter_column
# ---------------------------------------------------------------------------


class TestSKTFindQuarterColumn:
    def setup_method(self):
        self.parser = SKTExcelParser()

    def test_is_format_match(self):
        """Header "1Q24" at col index 1 should be found for year=2024, quarter=1."""
        ws = _make_ws_with_header(["Item", "1Q24", "2Q24", "3Q24", "4Q24"])
        result = self.parser._find_quarter_column(ws, header_row=1, year=2024, quarter=1)
        assert result == 1  # 0-indexed column 1

    def test_is_format_q3(self):
        ws = _make_ws_with_header(["Item", "1Q25", "2Q25", "3Q25", "4Q25"])
        result = self.parser._find_quarter_column(ws, header_row=1, year=2025, quarter=3)
        assert result == 3

    def test_bs_format_q1_match(self):
        """BS header "3-31-24" maps to Q1 2024."""
        ws = _make_ws_with_header(["Item", "3-31-24", "6-30-24"])
        result = self.parser._find_quarter_column(ws, header_row=1, year=2024, quarter=1)
        assert result == 1

    def test_bs_format_q4_match(self):
        """BS header "12-31-25" maps to Q4 2025."""
        ws = _make_ws_with_header(["Item", "3-31-25", "6-30-25", "9-30-25", "12-31-25"])
        result = self.parser._find_quarter_column(ws, header_row=1, year=2025, quarter=4)
        assert result == 4

    def test_no_match_returns_none(self):
        ws = _make_ws_with_header(["Item", "1Q23", "2Q23"])
        result = self.parser._find_quarter_column(ws, header_row=1, year=2025, quarter=1)
        assert result is None

    def test_case_insensitive_match(self):
        """Header comparison should be case-insensitive."""
        ws = _make_ws_with_header(["Item", "1q24"])
        result = self.parser._find_quarter_column(ws, header_row=1, year=2024, quarter=1)
        assert result == 1


# ---------------------------------------------------------------------------
# KTExcelParser._scan_sheet (in-memory workbook)
# ---------------------------------------------------------------------------


class TestKTScanSheet:
    def setup_method(self):
        self.parser = KTExcelParser()

    def test_finds_revenue_row(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        # Row: label in col A, Q1 value in col B
        ws.cell(row=1, column=1, value="매출액")
        ws.cell(row=1, column=2, value=500000)

        metrics: dict = {}
        issues: list = []
        found = self.parser._scan_sheet(ws, quarter=1, metrics_dict=metrics, issues=issues)

        assert found is True
        assert metrics.get("revenue") == 500000

    def test_unrecognised_label_ignored(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="알수없는항목")
        ws.cell(row=1, column=2, value=12345)

        metrics: dict = {}
        issues: list = []
        found = self.parser._scan_sheet(ws, quarter=1, metrics_dict=metrics, issues=issues)

        assert found is False
        assert metrics == {}

    def test_non_numeric_value_creates_issue(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="매출액")
        ws.cell(row=1, column=2, value="N/A")

        metrics: dict = {}
        issues: list = []
        self.parser._scan_sheet(ws, quarter=1, metrics_dict=metrics, issues=issues)

        assert len(issues) == 1
        assert issues[0].issue_type == "format_mismatch"
